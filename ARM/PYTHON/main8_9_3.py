import time
import serial
import serial.tools.list_ports
import threading
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def send_temperature():
    sendData = f"TEMPERATURE=?\n"
    my_serial.write( sendData.encode() )

def send_humidity():
    sendData = f"HUMIDITY=?\n"
    my_serial.write( sendData.encode() )

serial_receive_data = ""
def serial_read_thread():
    global serial_receive_data
    while True:
        read_data = my_serial.readline()
        serial_receive_data = read_data.decode()

def send_temp_humi():
    t2 = threading.Timer(1, send_temp_humi)
    t2.daemon = True
    t2.start()
    send_temperature()
    time.sleep(0.2)
    send_humidity()
    time.sleep(0.2)

def main():
    try:
        send_temp_humi()
        global serial_receive_data
        temperature = 0
        humidity = 0
        while True:
            if "TEMPERATURE=" in serial_receive_data:
                temperature = float(serial_receive_data[12:])
                serial_receive_data = ""
            elif "HUMIDITY=" in serial_receive_data:
                humidity = float(serial_receive_data[9:])
                serial_receive_data = ""
            
            if temperature !=0 and humidity !=0 :
                
                save_file_path = r"온습도데이터.xlsx"

                try:
                    wb = load_workbook(save_file_path, data_only=True)
                    sheet  = wb.active
                except:
                    wb = Workbook()
                    sheet = wb.active

                now = datetime.datetime.now()
                not_day_time = now.strftime('%Y-%m-%d %H:%M:%S')
                
                sheet.append([not_day_time,temperature,humidity])
                
                wb.save(save_file_path)
                
                print("날짜시간:",not_day_time,"온도:",temperature,"습도:",humidity)
                temperature = 0
                humidity = 0
                
    except KeyboardInterrupt:
        pass

if __name__ == '__main__':
    
    ports = list(serial.tools.list_ports.comports())
    for p in ports:
        if 'CH340' in p.description:
            print(f"{p} 포트에 연결하였습니다.")
            my_serial = serial.Serial(p.device, baudrate=9600, timeout=1.0)
            time.sleep(2.0)

    t1 = threading.Thread(target=serial_read_thread)
    t1.daemon = True
    t1.start()
    
    main()
    
    my_serial.close()